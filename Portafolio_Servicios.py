# ======================================================================================
# PORTAFOLIO DE SERVICIOS ESTRATÉGICOS: GM-DATOVATE
# VERSIÓN: 6.8 (Edición "Corrección TypeError en st.button")
# CORRECCIÓN CRÍTICA:
# 1. (BUG FIX 6.8) Corregido 'TypeError' en 'st.button' (Demo Catálogo).
#    Se cambió el parámetro incorrecto 'use_column_width' por
#    el parámetro correcto 'use_container_width'.
# 2. (BUG FIX 6.7) Corregido 'AttributeError' en 'pdf.output()'.
#    El método '.output()' en fpdf2 devuelve 'bytes', no 'str'.
# 3. (BUG FIX 6.6) Eliminado 'st.rerun()' explícito después de 'st.toast()'
#    en 'render_pagina_comercial' para prevenir el error 'NotFoundError: removeChild'.
# 4. (BUG FIX 6.5) Actualizados todos los parámetros deprecados.
# 5. (BUG FIX 6.4) Blindado DemoPDF.add_table() para tipos de fecha (previo fix).
#
# NOTA DE ENTORNO: Esta app requiere 'kaleido' en requirements.txt
# Y las dependencias de sistema en 'packages.txt' para Streamlit Cloud
# (ver notas de la solución anterior).
# ======================================================================================

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import datetime as dt # Importación explícita para evitar conflictos y mejorar claridad
import io
import urllib.parse
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import base64
from streamlit_drawable_canvas import st_canvas
import time

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(
    page_title="GM-DATOVATE | Ecosistemas de Inteligencia Empresarial",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="collapsed" # La barra lateral se oculta por defecto
)

# --- PALETA DE COLORES Y ESTILOS ---
COLOR_PRIMARIO = "#0D3B66"      # Azul profundo (Confianza, Inteligencia)
COLOR_SECUNDARIO = "#1A73E8"    # Azul brillante (Tecnología, Innovación)
COLOR_ACENTO_ROJO = "#F94144"       # Rojo vivo (Acción, Alerta)
COLOR_ACENTO_VERDE = "#43AA8B"      # Verde (Finanzas, Crecimiento)
COLOR_ACENTO_NARANJA = "#F8961E" # Naranja (Energía, CTA)
COLOR_FONDO = "#FFFFFF"     # Fondo Blanco Limpio
COLOR_FONDO_SECUNDARIO = "#F7F9FC" # Fondo gris muy claro para secciones
COLOR_TEXTO = "#2F2F2F"
COLOR_TEXTO_SECUNDARIO = "#555555"

# --- IMÁGENES EMBEBIDAS EN BASE64 (SVG Placeholders) ---
IMG_TEAM_DIEGO = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNTAiIGhlaWdodD0iMTUwIiB2aWV3Qm94PSIwIDAgMjQgMjQiIGZpbGw9IiNGRjdGOUZDIiBzdHJva2U9IiMwRDNCNjYiIHN0cm9rZS13aWR0aD0iMSI+PHJlY3Qgd2lkdGg9IjI0IiBoZWlnaHQ9IjI0IiBmaWxsPSIjRjdGOUZDIj48L3JlY3Q+PHBhdGggZD0iTTIwIDIxdi0yYTQgNCAwIDAgMC00LTRINWE0IDQgMCAwIDAtNCA0djJNNCA3YTYgNiAwIDEgMSAxMiAwIDYgNiAwIDAgMS0xMiAwWk0xNy41IDEyLjVMMTcgMTAuNWwyLTVoNGwxLjUgMyI+PC9wYXRoPjwvc3ZnPg=="
IMG_TEAM_PABLO = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNTAiIGhlaWdodD0iMTUwIiB2aWV3Qm94PSIwIDAgMjQgMjQiIGZpbGw9IiNGRjdGOUZDIiBzdHJva2U9IiMwRDNCNjYiIHN0cm9rZS13aWR0aD0iMSI+PHJlY3Qgd2lkdGg9IjI0IiBoZWlnaHQ9IjI0IiBmaWxsPSIjRjdGOUZDIj48L3JlY3Q+PHBhdGggZD0iTTIwIDIxdi0yYTQgNCAwIDAgMC00LTRINWE0IDQgMCAwIDAtNCA0djJNNCA3YTYgNiAwIDEgMSAxMiAwIDYgNiAwIDAgMS0xMiAwWk0xNyAxMGwxLjUgMS41TDYgMjFIMTciPjwvc3ZnPg=="
IMG_PROD_DISCO = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9IiM0MjQyNDIiIHN0cm9rZS13aWR0aD0iMiIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIiBzdHJva2UtdGluZWpvaW49InJvdW5kIj48Y2lyY2xlIGN4PSIxMiIgY3k9IjEyIiByPSI5Ij48L2NpcmNsZT48Y2lyY2xlIGN4PSIxMiIgY3k9IjEyIiByPSIzIj48L2NpcmNsZT48L3N2Zz4="
IMG_PROD_TORNILLO = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9IiM0MjQyNDIiIHN0cm9rZS13aWR0aD0iMS41IiBzdHJva2UtbGluZWNhcD0icm91bmQiIHN0cm9rZS1saW5lam9pbj0icm91bmQiPjxwYXRoIGQ9Ik0xOC4zNCA1LjY2YTIuMTIgMi4xMiAwIDAgMSAwIDNMNi4yMSAxOS44M2wtMi0yTDExLjM0IDguNjZhMi4xMiAyLjEyIDAgMCAxIDMtM3oiPjwvcGF0aD48bGluZSB4MT0iMyIgeTE9IjE0IiB4Mj0iMTAiIHkyPSIyMSI+PC9saW5lPjxsaW5lIHgxPSI3IiB5MT0iMTIiIHgyPSIxMiIgeTI9IjE3Ij48L2xpbmU+PGxpbmUgeDE9IjExIiB5MT0iOCIgeDI9IjE2IiB5Mj0iMTMiPjwvbGluZT48L3N2Zz4="
IMG_PROD_ELECTRODO = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9IiM0MjQyNDIiIHN0cm9rZS13aWR0aD0iMS41IiBzdHJva2UtbGluZWNhcD0icm91bmQiIHN0cm9rZS1saW5lam9pbj0icm91bmQiPjxwYXRoIGQ9Ik0yIDE4SDVhMyAzIDAgMCAwIDMtM1Y5YTMgMyAwIDAgMCAzLTNoM2EzIDMgMCAwIDAgMy0zVjIiPjwvcGF0aD48bGluZSB4MT0iMTIiIHkxPSI2IiB4Mj0iOCIgeTI9IjEwIj48L2xpbmU+PGxpbmUgeDE9IjgiIHkxPSIxNCIgeDI9IjEwIiB5Mj0iMTIiPjwvbGluZT48bGluZSB4MT0iMTYiIHkxPSI0IiB4Mj0iMTQiIHkyPSI2Ij48L2xpbmU+PGxpbmUgeDE9IjUiIHkxPSIyMiIgeDI9IjIiIHkyPSIxOCI+PC9saW5lPjxsaW5lIHgxPSI5IiB5MT0iMTgiIHgyPSI1IiB5Mj0iMTQiPjwvbGluZT48L3N2Zz4="
IMG_PROD_GAFA = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9IiM0MjQyNDIiIHN0cm9rZS13aWR0aD0iMS41IiBzdHJva2UtbGluZWNhcD0icm91bmQiIHN0cm9rZS1saW5lam9pbj0icm91bmQiPjxjaXJjbGUgY3g9IjYuNSIgY3k9IjE1LjUiIHI9IjQuNSI+PC9jaXJjbGU+PGNpcmNsZSBjeD0iMTcuNSIgY3k9IjE1LjUiIHI9IjQuNSI+PC9jaXJjbGU+PHBhdGggZD0iTTIgMTUuNWE0LjUgNC41IDAgMCAwIDQuNSA0LjVoMTFBNi41IDYuNSAwIDAgMCAyMiAxNS41Ij48L3BhdGg+PHBhdGggZD0iTTYuNSAxMS41YTQuNSAzMCAwIDEgMC05bTEuMiAzLjJsMS4zLTEuNW04LjUgNy4zbDEuNSAxLjUiPjwvcGF0aD48L3N2Zz4="
IMG_PROD_GUANTE = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9IiM0MjQyNDIiIHN0cm9rZS13aWR0aD0iMS41IiBzdHJva2UtbGluZWNhcD0icm91bmQiIHN0cm9rZS1saW5lam9pbj0icm91bmQiPjxwYXRoIGQ9Ik0yMiAxNGExMiAxMiAwIDAgMS04IDhIMWwtMy04VjZhNCA0IDAgMCAxIDQtNGg1LjVMMTAgNyI+PC9wYXRoPjxwYXRoIGQ9Ik0xMS41IDZhNC41IDQuNSAwIDEgMSAwIDlWNmEiPjwvcGF0aD48cGFhdGggZD0iJNMTYgNmE0IDQgMCAwIDEgMCA4VjYiPjwvcGF0aD48cGFhdGggZD0iJNTE5IDZhMyAzIDAgMCAxIDAgNlY2Ij48L3BhdGg+PC9zdmc+"

# --- INYECCIÓN DE CSS GLOBAL (PROFESIONAL) ---
st.markdown(f"""
<style>
    /* --- Ocultar elementos de Streamlit --- */
    #MainMenu {{display: none;}}
    footer {{display: none;}}
    [data-testid="stHeader"] {{display: none;}}
    [data-testid="stSidebar"] {{display: none;}}

    /* --- Reseteo y Fuentes --- */
    body {{
        font-family: 'Arial', sans-serif;
        color: {COLOR_TEXTO};
    }}

    /* --- Contenedor Principal --- */
    .main .block-container {{
        padding: 0rem;
    }}
    
    /* Forzar Fondo Claro */
    .stApp {{
        background-color: {COLOR_FONDO} !important;
    }}
    body {{
        background-color: {COLOR_FONDO} !important;
    }}

    /* --- Animaciones --- */
    @keyframes fadeIn {{
        from {{ opacity: 0; transform: translateY(20px); }}
        to {{ opacity: 1; transform: translateY(0); }}
    }}
    .section-container, .stTabs, .hero-container {{
        animation: fadeIn 0.8s ease-out;
    }}

    /* --- Clases de Diseño Personalizadas --- */
    .section-container {{
        padding: 4rem 2rem;
    }}
    .section-container-dark {{
        background-color: {COLOR_FONDO_SECUNDARIO};
        border-top: 1px solid #E0E0E0;
        border-bottom: 1px solid #E0E0E0;
    }}
    
    /* --- Héroe (Banner Principal) --- */
    .hero-container {{
        display: flex;
        align-items: center;
        justify-content: center;
        text-align: center;
        padding: 6rem 2rem;
        background: linear-gradient(135deg, {COLOR_PRIMARIO} 0%, {COLOR_SECUNDARIO} 100%);
        color: white;
        min-height: 70vh;
    }}
    .hero-content {{
        max-width: 800px;
        animation: fadeIn 1s ease-out;
    }}
    .hero-container h1 {{
        font-size: 3.8rem;
        font-weight: 700;
        color: white;
        margin: 0;
        line-height: 1.2;
    }}
    .hero-container h3 {{
        font-size: 1.6rem;
        font-weight: 400;
        color: {COLOR_ACENTO_NARANJA};
        margin-top: 1rem;
    }}
    .hero-container p {{
        font-size: 1.1rem;
        margin: 1.5rem 0;
        color: #E0E0E0;
    }}
    .hero-button {{
        display: inline-block;
        padding: 0.9rem 2.2rem;
        background-color: {COLOR_ACENTO_ROJO};
        color: white;
        font-weight: bold;
        text-decoration: none;
        border-radius: 8px;
        transition: all 0.3s;
        margin-top: 1.5rem;
        font-size: 1.1rem;
        border: 2px solid {COLOR_ACENTO_ROJO};
    }}
    .hero-button:hover {{
        background-color: #FFFFFF;
        color: {COLOR_ACENTO_ROJO};
        transform: scale(1.05);
    }}

    /* --- Tarjetas de Servicios y Equipo --- */
    .service-card, .team-card {{
        height: 100%;
        background-color: {COLOR_FONDO};
        border-radius: 10px;
        padding: 2rem;
        box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        border: 1px solid #E0E0E0;
        transition: all 0.3s ease;
    }}
    .service-card:hover, .team-card:hover {{
        transform: translateY(-10px);
        box-shadow: 0 12px 24px rgba(0,0,0,0.1);
    }}
    .service-card-icon {{
        font-size: 3.5rem;
        line-height: 1;
    }}
    .service-card h3 {{
        margin-top: 1rem;
        color: {COLOR_PRIMARIO};
        font-size: 1.4rem;
    }}
    .team-card img {{ 
        width: 150px;
        height: 150px;
        border-radius: 50%;
        border: 5px solid {COLOR_PRIMARIO};
        box-shadow: 0 4px 10px rgba(0,0,0,0.1);
        background-color: #f0f0f0;
    }}
    .team-card h3 {{ color: {COLOR_PRIMARIO}; margin-top: 1rem; }}

    /* --- Pestañas de Demo --- */
    .stTabs {{
        background-color: {COLOR_FONDO_SECUNDARIO};
        padding: 2rem 2rem 4rem 2rem;
        border-top: 1px solid #E0E0E0;
    }}
    .stTabs [data-baseweb="tab-list"] {{
        gap: 24px;
        padding-left: 1rem;
        padding-right: 1rem;
        border-bottom: 2px solid #E0E0E0;
    }}
    .stTabs [aria-selected="true"] {{
        border-bottom: 4px solid {COLOR_ACENTO_ROJO};
        color: {COLOR_PRIMARIO};
        font-weight: bold;
    }}
    
    /* --- Botones --- */
    .stButton>button {{
        border-radius: 8px;
        font-weight: bold;
        transition: all 0.3s;
        padding: 0.75rem 1.5rem;
    }}
    .stButton>button[kind="primary"] {{
        border: 2px solid {COLOR_PRIMARIO};
        background-color: {COLOR_PRIMARIO};
        color: white;
    }}
    .stButton>button[kind="primary"]:hover {{
        background-color: {COLOR_SECUNDARIO};
        border-color: {COLOR_SECUNDARIO};
        color: white;
    }}

    /* --- Media Query para móviles (Responsivo) --- */
    @media (max-width: 768px) {{
        .hero-container h1 {{ font-size: 2.2rem; }}
        .section-container {{ padding: 2rem 1.5rem; }}
        
        /* Forzar apilado de st.columns en móviles */
        .stTabs [data-testid="stHorizontalBlock"],
        .contact-form-container [data-testid="stHorizontalBlock"] {{
            flex-direction: column !important;
            flex-wrap: nowrap !important;
        }}
        .stTabs [data-testid="stHorizontalBlock"] > div[data-testid="stVerticalBlock"],
        .contact-form-container [data-testid="stHorizontalBlock"] > div[data-testid="stVerticalBlock"] {{
            width: 100% !important;
            margin-bottom: 1rem;
        }}
    }}
</style>
""", unsafe_allow_html=True)

# --- FUNCIÓN AUXILIAR PARA RENDERIZAR KPIs PERSONALIZADOS ---
def render_kpi(title, value, delta=None, delta_color="normal", card_color=""):
    """Renderiza una tarjeta de KPI con CSS personalizado."""
    delta_html = ""
    if delta:
        if delta_color == "inverse":
            color = COLOR_ACENTO_VERDE
            arrow = "↑" if delta.startswith('+') or not delta.startswith('-') else "↓"
        elif delta_color == "normal":
            color = COLOR_ACENTO_ROJO
            arrow = "↓" if delta.startswith('-') else "↑"
        else: # off
            color = COLOR_TEXTO_SECUNDARIO
            arrow = ""
        
        delta_html = f'<div class="kpi-delta" style="font-size: 0.9rem; font-weight: 600; color: {color};">{arrow} {delta}</div>'
    
    if card_color == "red":
        border_color = COLOR_ACENTO_ROJO
    elif card_color == "green":
        border_color = COLOR_ACENTO_VERDE
    else:
        border_color = COLOR_PRIMARIO

    st.markdown(f"""
    <div style="background-color: #FFFFFF; border: 1px solid #E0E0E0; border-left: 5px solid {border_color}; border-radius: 8px; padding: 1.2rem 1.5rem; margin-bottom: 1rem; box-shadow: 0 2px 8px rgba(0,0,0,0.05);">
        <div style="font-size: 0.9rem; color: {COLOR_TEXTO_SECUNDARIO}; font-weight: 600; margin-bottom: 0.5rem; text-transform: uppercase;">{title}</div>
        <div style="font-size: 2rem; font-weight: 700; color: {COLOR_PRIMARIO}; line-height: 1;">{value}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)

# ======================================================================================
# --- DATOS DE EJEMPLO PARA LAS DEMOS (CACHEABLES) ---
# ======================================================================================
@st.cache_data
def get_sample_data():
    """Crea y cachea todos los DataFrames de ejemplo para las demos."""
    data = {}
    
    # --- Datos Comerciales ---
    vendedores = ['DIEGO GARCIA', 'ANGELA CONTRERAS', 'PABLO MAFLA', 'MARY LUZ TREJOS']
    regiones = ['EJE CAFETERO', 'ANTIOQUIA', 'EJE CAFETERO', 'ANTIOQUIA']
    data['ventas_vendedor'] = pd.DataFrame({
        'Vendedor': vendedores,
        'Region': regiones,
        'Ventas ($)': [120_000_000, 95_000_000, 88_000_000, 75_000_000],
        'Meta ($)': [110_000_000, 100_000_000, 90_000_000, 80_000_000],
    })
    data['ventas_vendedor']['Avance (%)'] = (data['ventas_vendedor']['Ventas ($)'] / data['ventas_vendedor']['Meta ($)']) * 100

    # Datos de series de tiempo para ventas
    np.random.seed(42)
    fechas = pd.to_datetime([datetime.now().date() - timedelta(days=x) for x in range(30)])
    ventas_dia = np.random.randint(8_000_000, 15_000_000, size=30)
    data['ventas_dia'] = pd.DataFrame({'Fecha': fechas, 'Ventas ($)': ventas_dia}).sort_values('Fecha')
    
    data['ventas_categoria'] = pd.DataFrame({
        'Categoria': ['HERRAMIENTA ELÉCTRICA', 'TORNILLERÍA', 'SEGURIDAD INDUSTRIAL', 'ABRASIVOS'],
        'Ventas ($)': [110_000_000, 90_000_000, 75_000_000, 103_000_000]
    })

    # --- Datos de Cotizador ---
    data['catalogo_productos'] = pd.DataFrame({
        'Referencia': ['A-101', 'B-202', 'C-303', 'D-404', 'E-505'],
        'Producto': ['Disco Corte 4-1/2"', 'Tornillo Drywall 6x1', 'Electrodo 6013', 'Gafa de Seguridad', 'Guante Vaqueta'],
        'Vlr. Unitario': [1800, 150, 800, 4500, 7000],
        'Stock': [500, 15000, 800, 120, 300],
        'ImagenURL': [
            IMG_PROD_DISCO, IMG_PROD_TORNILLO, IMG_PROD_ELECTRODO,
            IMG_PROD_GAFA, IMG_PROD_GUANTE
        ]
    })

    # --- Datos Operaciones (Abastecimiento) ---
    data['sugerencia_abastecimiento'] = pd.DataFrame({
        'SKU': ['A-101', 'B-202', 'C-303', 'D-404'],
        'Producto': ['Disco Corte 4-1/2"', 'Tornillo Drywall 6x1', 'Electrodo 6013', 'Gafa de Seguridad'],
        'Stock (Total)': [500, 15000, 800, 120],
        'Stock Tránsito': [0, 5000, 0, 100],
        'Necesidad Real': [200, 10000, 500, 150],
        'Sugerencia Traslado': [0, 0, 0, 0],
        'Sugerencia Compra': [0, 0, 0, 0]
    })

    # --- Datos Financieros (Cartera) ---
    data['cartera_antiguedad'] = pd.DataFrame({
        'Rango': ['Al día', '1-15 días', '16-30 días', '31-60 días', 'Más de 60 días'],
        'Valor ($)': [250_000_000, 80_000_000, 45_000_000, 25_000_000, 70_000_000],
        'Color': [COLOR_ACENTO_VERDE, '#FBC02D', '#F57C00', COLOR_ACENTO_ROJO, '#a71919']
    })
    data['cartera_detalle'] = pd.DataFrame({
        'Cliente': ['Cliente A', 'Cliente B (Vencido)', 'Cliente B (Vencido)', 'Cliente C', 'Cliente D (Vencido)'],
        'NIT': ['800.123.456-1', '800.987.654-2', '800.987.654-2', '800.222.333-4', '900.555.444-5'],
        'Factura': ['FV-1001', 'FV-901', 'FV-902', 'FV-1002', 'FV-850'],
        'Fecha Vencimiento': [datetime.now().date() + timedelta(days=15), datetime.now().date() - timedelta(days=45), datetime.now().date() - timedelta(days=12), datetime.now().date() + timedelta(days=30), datetime.now().date() - timedelta(days=70)],
        'Días Vencido': [-15, 45, 12, -30, 70],
        'Monto': [2_500_000, 1_200_000, 850_000, 3_100_000, 4_500_000],
        'Email': ['pagos@clientea.com', 'pagos@clienteb.com', 'pagos@clienteb.com', 'pagos@clientec.com', 'pagos@cliented.com'],
        'Teléfono': ['573001234567', '573019876543', '573019876543', '573022223333', '573035554444']
    })
    
    # Datos para Integración de Riesgo
    data['covinoc_subir'] = pd.DataFrame({
        'Factura': ['FV-1001', 'FV-1002', 'FV-1003'], 'Cliente': ['Cliente A', 'Cliente B', 'Cliente C'],
        'Días Vencido': [35, 40, 28], 'Monto': [1_200_000, 850_000, 2_100_000]
    })
    data['covinoc_exonerar'] = pd.DataFrame({
        'Factura': ['FV-901', 'FV-902'], 'Cliente': ['Cliente D', 'Cliente E'],
        'Estado Covinoc': ['Pendiente', 'Pendiente'], 'Estado Cartera': ['PAGADA', 'PAGADA']
    })
    
    return data

SAMPLE_DATA = get_sample_data()


# --- INICIALIZACIÓN DE SESSION STATE ---
if 'cart' not in st.session_state:
    st.session_state.cart = pd.DataFrame(columns=['Referencia', 'Producto', 'Cantidad', 'Vlr. Unitario', 'Total'])
if 'chat_messages' not in st.session_state:
    st.session_state.chat_messages = []
if 'otp_sent' not in st.session_state:
    st.session_state.otp_sent = False
if 'otp_code' not in st.session_state:
    st.session_state.otp_code = ""
if 'cuadre_data' not in st.session_state:
    st.session_state.cuadre_data = pd.DataFrame({
        'Tipo': ['Tarjetas', 'Consignaciones', 'Gastos', 'Efectivo Entregado'],
        'Valor': [2500000, 1500000, 200000, 800000]
    })

# ======================================================================================
# --- CLASES DE GENERACIÓN DE DOCUMENTOS (PDF Y EXCEL) ---
# ======================================================================================

class DemoPDF(FPDF):
    """Clase personalizada para generar PDFs profesionales."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title = "Documento de Demostración"

    def header(self):
        # Fondo del encabezado con color primario (convertir HEX a RGB)
        self.set_fill_color(int(COLOR_PRIMARIO[1:3], 16), int(COLOR_PRIMARIO[3:5], 16), int(COLOR_PRIMARIO[5:7], 16))
        self.rect(0, 0, self.w, 30, 'F')
        
        self.set_font('Arial', 'B', 20)
        self.set_text_color(255, 255, 255)
        self.set_xy(10, 10)
        self.cell(0, 10, 'GM-DATOVATE', 0, 0, 'L')

        # Título del documento (convertir HEX a RGB)
        hex_acento = COLOR_ACENTO_NARANJA.replace("#", "")
        self.set_text_color(int(hex_acento[0:2], 16), int(hex_acento[2:4], 16), int(hex_acento[4:6], 16))
        self.set_xy(10, 18)
        self.set_font('Arial', 'B', 15)
        self.cell(0, 10, self.title, 0, 1, 'R')
        self.ln(15)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, 'Ecosistema de Inteligencia GM-DATOVATE | Página %s' % self.page_no(), 0, 0, 'C')

    def chapter_title(self, title):
        self.set_font('Arial', 'B', 12)
        # Convertir HEX a RGB para fondo secundario
        hex_fondo = COLOR_FONDO_SECUNDARIO.replace("#", "")
        self.set_fill_color(int(hex_fondo[0:2], 16), int(hex_fondo[2:4], 16), int(hex_fondo[4:6], 16))
        # Convertir HEX a RGB para color primario
        hex_primario = COLOR_PRIMARIO.replace("#", "")
        self.set_text_color(int(hex_primario[0:2], 16), int(hex_primario[2:4], 16), int(hex_primario[4:6], 16))
        
        self.cell(0, 10, f" {title}", 0, 1, 'L', fill=True)
        self.ln(4)
        
    def chapter_body(self, body):
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        self.multi_cell(0, 5, body)
        self.ln()
        
    def add_chart(self, chart_image, width_percent=0.8):
        """Añade una imagen de gráfico (desde bytes) al PDF."""
        if chart_image:
            try:
                img_width = self.w * width_percent
                img_x = (self.w - img_width) / 2
                self.image(chart_image, x=img_x, w=img_width, type='PNG')
            except Exception as e:
                self.set_text_color(255, 0, 0)
                self.cell(0, 5, f"Error al renderizar el gráfico: {e}")
                self.set_text_color(0,0,0)
                self.ln(5)

    def add_table(self, df):
        if df.empty:
            # Uso correcto de RGB para texto secundario
            hex_texto_sec = COLOR_TEXTO_SECUNDARIO.replace("#", "")
            r = int(hex_texto_sec[0:2], 16)
            g = int(hex_texto_sec[2:4], 16)
            b = int(hex_texto_sec[4:6], 16)
            self.set_text_color(r, g, b) 
            
            self.set_font('Arial', 'I', 10)
            self.cell(0, 10, "No hay datos para mostrar en la tabla.", 0, 1)
            self.set_text_color(0, 0, 0)
            return

        self.set_font('Arial', 'B', 9)
        # Convertir HEX a RGB para cabecera de tabla
        hex_sec = COLOR_SECUNDARIO.replace("#", "")
        self.set_fill_color(int(hex_sec[0:2], 16), int(hex_sec[2:4], 16), int(hex_sec[4:6], 16))
        self.set_text_color(255, 255, 255)
        
        # Lógica de cálculo de ancho de columna
        page_width = self.w - 2 * self.l_margin
        num_cols = len(df.columns)
        
        col_widths = {}
        for col in df.columns:
            if 'Producto' in str(col) or 'Cliente' in str(col):
                col_widths[col] = page_width * 0.3
            elif 'Monto' in str(col) or 'Valor' in str(col) or 'Total' in str(col):
                col_widths[col] = page_width * 0.2
        
        fixed_width = sum(col_widths.values())
        fixed_cols = len(col_widths)
        non_fixed_cols = num_cols - fixed_cols
        
        dynamic_width = 0
        if non_fixed_cols > 0:
            dynamic_width = (page_width - fixed_width) / non_fixed_cols
            
        final_widths = []
        for col in df.columns:
            final_widths.append(col_widths.get(col, dynamic_width))

        # Renderizar cabecera
        for i, header in enumerate(df.columns):
            self.cell(final_widths[i], 7, str(header), 1, 0, 'C', fill=True)
        self.ln()

        # Renderizar cuerpo
        self.set_font('Arial', '', 9)
        self.set_text_color(0, 0, 0)
        self.set_fill_color(245, 245, 245)
        fill = False
        for _, row in df.iterrows():
            for i, item in enumerate(row):
                align = 'L'
                item_str = str(item)
                
                # Formato de números y fechas
                if isinstance(item, (int, float, np.integer, np.floating)):
                    align = 'R'
                    if "Monto" in df.columns[i] or "Valor" in df.columns[i] or "Total" in df.columns[i] or "Vlr. Unitario" in df.columns[i] or "Ventas" in df.columns[i]:
                        item_str = f"${item:,.0f}"
                    elif "Avance" in df.columns[i]:
                        item_str = f"{item:,.1f}%"
                    else:
                        item_str = f"{item:,.0f}"
                
                # --- VERIFICACIÓN DE TIPO DE FECHA BLINDADA ---
                # Verifica instancias de tipos de fecha/tiempo de Python y Pandas.
                elif isinstance(item, (dt.datetime, pd.Timestamp, dt.date)):
                    try:
                        # Convertir a pd.Timestamp de forma segura para usar strftime
                        timestamp = pd.to_datetime(item)
                        item_str = timestamp.strftime('%Y-%m-%d')
                        align = 'C'
                    except Exception:
                        # Fallback a string si la conversión falla
                        item_str = str(item)
                        align = 'L'
                        
                self.cell(final_widths[i], 6, item_str, 1, 0, align, fill=fill)
            self.ln()
            fill = not fill

@st.cache_data
def generar_demo_pdf(df, title, intro_text, chart_fig=None):
    """Genera un PDF genérico (MEJORADO con gráfico opcional)."""
    pdf = DemoPDF()
    pdf.title = title
    pdf.add_page()
    pdf.chapter_title("Resumen del Reporte")
    pdf.chapter_body(intro_text)
    
    if chart_fig:
        pdf.chapter_title("Análisis Visual")
        img_bytes = io.BytesIO(chart_fig.to_image(format="png", scale=2))
        pdf.add_chart(img_bytes)
        pdf.ln(5)

    pdf.chapter_title("Datos Detallados")
    pdf.add_table(df)
    
    # --- CORRECCIÓN (BUG FIX 6.7) ---
    # .output() sin 'dest' (o con dest='S') en fpdf2 devuelve bytes.
    # No se debe llamar a .encode() sobre un objeto bytes.
    return pdf.output()

@st.cache_data
def generar_demo_pdf_cartera(df, cliente_info, chart_fig):
    """Genera un PDF de estado de cuenta (MEJORADO con gráfico)."""
    pdf = DemoPDF()
    pdf.title = "Estado de Cuenta"
    pdf.add_page()
    
    pdf.chapter_title("Información del Cliente")
    # ... Lógica de info de cliente (usa set_text_color(r, g, b) ya corregido en los estilos) ...
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(0, 0, 0) # Asegurar color de texto
    pdf.cell(40, 6, "Cliente:")
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, cliente_info['Cliente'])
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 6, "NIT:")
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, cliente_info['NIT'])
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 6, "Teléfono:")
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, cliente_info['Teléfono'])
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 6, "Email:")
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, cliente_info['Email'])
    pdf.ln(10)

    pdf.chapter_title("Detalle de Facturas Vencidas")
    pdf.add_table(df[['Factura', 'Fecha Vencimiento', 'Días Vencido', 'Monto']])
    pdf.ln(5)
    
    total_vencido = df['Monto'].sum()
    pdf.set_font('Arial', 'B', 12)
    # Convertir HEX a RGB para acento rojo
    hex_rojo = COLOR_ACENTO_ROJO.replace("#", "")
    pdf.set_fill_color(int(hex_rojo[0:2], 16), int(hex_rojo[2:4], 16), int(hex_rojo[4:6], 16))
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, f"TOTAL VENCIDO: ${total_vencido:,.0f}", 1, 1, 'C', fill=True)
    pdf.ln(10)

    pdf.chapter_title("Composición de la Deuda (Global)")
    if chart_fig:
        # Esta es la línea que falla si 'kaleido' no está instalado
        # o si faltan las dependencias del sistema (ver packages.txt)
        img_bytes = io.BytesIO(chart_fig.to_image(format="png", scale=2))
        pdf.add_chart(img_bytes)
        
    # --- CORRECCIÓN (BUG FIX 6.7) ---
    # .output() sin 'dest' (o con dest='S') en fpdf2 devuelve bytes.
    # No se debe llamar a .encode() sobre un objeto bytes.
    return pdf.output()

@st.cache_data
def generar_demo_excel(df_dict):
    """
    Genera un archivo Excel con múltiples hojas, formato profesional y hoja de resumen.
    (BLINDADO: Bug de TypeError al verificar tipos de fecha en DataFrames vacíos).
    """
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Estilos base
            title_font = Font(size=18, bold=True, color=COLOR_PRIMARIO.replace("#",""))
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color=COLOR_PRIMARIO.replace("#",""), fill_type="solid")
            kpi_title_font = Font(bold=True, color=COLOR_TEXTO_SECUNDARIO.replace("#",""))
            kpi_value_font = Font(size=14, bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            currency_format = '$ #,##0'
            date_format = 'dd/mm/yyyy'

            # 1. Hoja de Resumen (si hay datos)
            ws_resumen = writer.book.create_sheet("Resumen Ejecutivo", 0)
            ws_resumen.merge_cells('B2:E2')
            title_cell = ws_resumen['B2']
            title_cell.value = "Resumen Ejecutivo del Reporte"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center')
            
            ws_resumen.column_dimensions['B'].width = 30
            ws_resumen.column_dimensions['C'].width = 25

            # KPIs de ejemplo
            kpis = {
                "Total Facturas a Subir": (df_dict.get("1_Facturas_a_Subir", pd.DataFrame())['Monto'].sum() if "1_Facturas_a_Subir" in df_dict and not df_dict.get("1_Facturas_a_Subir").empty else 0),
                "Total Facturas a Exonerar": (df_dict.get("2_Facturas_a_Exonerar", pd.DataFrame())['Monto'].sum() if "2_Facturas_a_Exonerar" in df_dict and not df_dict.get("2_Facturas_a_Exonerar").empty else 0),
                "Total Orden de Compra": (df_dict.get("Orden_de_Compra", pd.DataFrame())['Sugerencia Compra'].sum() if "Orden_de_Compra" in df_dict and 'Sugerencia Compra' in df_dict.get("Orden_de_Compra").columns and not df_dict.get("Orden_de_Compra").empty else 0)
            }
            
            row = 5
            for k, v in kpis.items():
                if v and v > 0: # Solo mostrar KPIs relevantes
                    cell_title = ws_resumen.cell(row=row, column=2, value=k)
                    cell_value = ws_resumen.cell(row=row, column=3, value=v)
                    cell_title.font = kpi_title_font
                    cell_value.font = kpi_value_font
                    cell_value.number_format = currency_format
                    cell_title.border = thin_border
                    cell_value.border = thin_border
                    row += 1

            # 2. Generar Hojas de Detalle
            for sheet_name, df in df_dict.items():
                if df.empty:
                    continue
                
                # Escribir DataFrame en la hoja
                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
                ws = writer.sheets[sheet_name]
                
                for i, col in enumerate(df.columns, 1):
                    col_letter = get_column_letter(i)
                    cell = ws.cell(row=2, column=i)
                    
                    # Aplicar estilo de cabecera
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Autoajustar ancho (mejorado)
                    try:
                        max_len_col = df[col].astype(str).map(len).max() if not df[col].empty and df[col].notna().any() else 0
                        max_len = max(len(str(col)), max_len_col)
                        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
                    except Exception:
                        ws.column_dimensions[col_letter].width = 20 # Fallback
                    
                    # Aplicar formatos de columna (BLINDADO)
                    is_date_col = False
                    
                    if pd.api.types.is_datetime64_any_dtype(df[col]):
                        is_date_col = True
                    
                    elif df[col].dtype == 'object':
                        # Verifica de forma segura si la columna de objetos contiene fechas
                        if not df.empty and df[col].notna().any():
                            try:
                                # Acceder de forma segura al primer elemento no nulo
                                first_non_nan = df[col].dropna().iloc[0]
                                if isinstance(first_non_nan, (datetime, dt.date, pd.Timestamp)):
                                    is_date_col = True
                            except IndexError:
                                pass
                            except Exception:
                                pass

                    if is_date_col:
                        for c in ws[col_letter][2:]:
                            c.number_format = date_format
                            
                    if 'Monto' in col or 'Valor' in col or 'Total' in col or 'Vlr. Unitario' in col or 'Ventas' in col:
                        for c in ws[col_letter][2:]:
                            c.number_format = currency_format

                # Título
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = sheet_name.replace("_", " ")
                title_cell.font = Font(size=16, bold=True, color=COLOR_PRIMARIO.replace("#",""))
                title_cell.alignment = Alignment(horizontal='center')
    
    except Exception as e:
        st.error(f"Error fatal en la generación de Excel: {e}")
        return None 
        
    return output.getvalue()


# ======================================================================================
# --- FUNCIONES DE RENDERIZADO DE PÁGINAS ---
# ======================================================================================

def render_pagina_inicio():
    """Renderiza la página de bienvenida, el pitch de valor y el equipo."""
    
    # --- Sección Héroe ---
    with st.container():
        st.markdown(f"""
        <div class="hero-container">
            <div class="hero-content">
                <h1>El Sistema Operativo de su Empresa, Impulsado por Datos.</h1>
                <h3>Deje de reaccionar. Empiece a predecir.</h3>
                <p>
                    Somos GM-DATOVATE. No solo creamos dashboards; construimos 
                    <strong>Ecosistemas de Inteligencia Empresarial</strong>.
                    Conectamos sus Ventas, Finanzas, Operaciones e IA en un solo cerebro digital 
                    que le dice qué hacer, cuándo hacerlo y por qué.
                </p>
                <a href="#contacto" class="hero-button">
                    Solicitar Consulta Estratégica
                </a>
            </div>
        </div>
        """, unsafe_allow_html=True)


    # --- Sección de Servicios/Pilares ---
    with st.container():
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        st.markdown("<h2 style='text-align: center; border: none; margin-bottom: 3rem;'>Un Ecosistema, Cuatro Pilares de Valor</h2>", unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="service-card">
                <div class="service-card-icon" style="color: {COLOR_ACENTO_ROJO};">🧠</div>
                <h3>Inteligencia Comercial</h3>
                <p style="color: {COLOR_TEXTO_SECUNDARIO};">
                    Usamos IA para analizar su historial y decirle a quién llamar, 
                    qué producto ofrecer (venta cruzada) y qué clientes están en riesgo de irse.
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="service-card">
                <div class="service-card-icon" style="color: {COLOR_ACENTO_VERDE};">🏭</div>
                <h3>Operaciones y Logística</h3>
                <p style="color: {COLOR_TEXTO_SECUNDARIO};">
                    Automatizamos su cadena de suministro. Desde el abastecimiento inteligente 
                    hasta el conteo físico en bodega con apps móviles y sincronización con su ERP.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
        with col3:
            st.markdown(f"""
            <div class="service-card">
                <div class="service-card-icon" style="color: {COLOR_SECUNDARIO};">🏦</div>
                <h3>Finanzas y Tesorería</h3>
                <p style="color: {COLOR_TEXTO_SECUNDARIO};">
                    Digitalizamos su flujo de caja. Automatizamos cuadres, recibos de caja, 
                    gestión de cartera y generamos los archivos planos para su ERP.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
        with col4:
            st.markdown(f"""
            <div class="service-card">
                <div class="service-card-icon" style="color: {COLOR_ACENTO_NARANJA};">🤖</div>
                <h3>Integración y Automatización</h3>
                <p style="color: {COLOR_TEXTO_SECUNDARIO};">
                    Unimos todos los sistemas: vinculación de clientes con firma digital, gestión de riesgo 
                    con agencias externas (Covinoc) y un Agente IA en WhatsApp 24/7.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
        st.markdown('</div>', unsafe_allow_html=True)

    # --- Sección "Quiénes Somos" ---
    with st.container():
        st.markdown('<div class="section-container section-container-dark">', unsafe_allow_html=True)
        st.markdown("<h2 style='text-align: center; border: none; margin-bottom: 3rem;'>Nuestros Arquitectos de Soluciones</h2>", unsafe_allow_html=True)
        
        col_space1, col_team1, col_team2, col_space2 = st.columns([1, 2, 2, 1])
        
        with col_team1:
            st.markdown(f"""
            <div class="team-card">
                <img src="{IMG_TEAM_DIEGO}" alt="Foto de Diego Garcia">
                <h3>Diego Mauricio García</h3>
                <p>Arquitecto de Datos y Desarrollador Líder</p>
                <span style="font-size: 0.9rem; color: {COLOR_TEXTO_SECUNDARIO};">
                    Especialista en la construcción de ecosistemas de datos complejos, 
                    transformando procesos manuales en flujos de trabajo automatizados 
                    e inteligentes.
                </span>
            </div>
            """, unsafe_allow_html=True)
        
        with col_team2:
            st.markdown(f"""
            <div class="team-card">
                <img src="{IMG_TEAM_PABLO}" alt="Foto de Pablo Mafla">
                <h3>Pablo Cesar Mafla</h3>
                <p>Estratega Comercial y de Negocios</p>
                <span style="font-size: 0.9rem; color: {COLOR_TEXTO_SECUNDARIO};">
                    Experto en alinear la tecnología con los objetivos de negocio. 
                    Traduce las necesidades del mercado en herramientas de datos que 
                    impulsan el crecimiento y la rentabilidad.
                </span>
            </div>
            """, unsafe_allow_html=True)
            
        st.markdown('</div>', unsafe_allow_html=True)

def render_pagina_comercial():
    """Demo de la Suite de Inteligencia Comercial."""
    
    st.markdown(f"<h2 style='color: {COLOR_PRIMARIO};'>🧠 Inteligencia Comercial</h2>", unsafe_allow_html=True)
    st.markdown("Deje que sus datos le digan cómo vender más. Automatizamos la prospección, la cotización y el análisis de rendimiento.")
    st.divider()
    
    tab1, tab2, tab3 = st.tabs([
        "Demo 1: Dashboard de BI Gerencial (Interactivo)",
        "Demo 2: Asistente Proactivo (Simulación IA)",
        "Demo 3: Catálogo y Cotizador (Interactivo)"
    ])
    
    # --- DEMO 1: BI GERENCIAL (INTERACTIVO) ---
    with tab1:
        st.subheader("Dashboard de BI Gerencial (En Tiempo Real)")
        
        with st.container(border=True):
            df_ventas = SAMPLE_DATA['ventas_vendedor']
            
            st.markdown("#####  Filtros del Dashboard")
            filtro_c1, filtro_c2 = st.columns(2)
            vendedores_filtro = filtro_c1.multiselect(
                "Filtrar por Vendedor:", 
                options=df_ventas['Vendedor'].unique(),
                default=df_ventas['Vendedor'].unique()
            )
            regiones_filtro = filtro_c2.multiselect(
                "Filtrar por Región:",
                options=df_ventas['Region'].unique(),
                default=df_ventas['Region'].unique()
            )
            
            df_filtrada = df_ventas[
                (df_ventas['Vendedor'].isin(vendedores_filtro)) &
                (df_ventas['Region'].isin(regiones_filtro))
            ]
            
            if df_filtrada.empty:
                st.warning("No hay datos para los filtros seleccionados.")
                
            else:
                st.markdown("##### KPIs Generales")
                total_ventas = df_filtrada['Ventas ($)'].sum()
                total_meta = df_filtrada['Meta ($)'].sum()
                avg_avance = (total_ventas / total_meta) * 100 if total_meta > 0 else 0
                delta_avance = avg_avance - 100

                col1, col2, col3 = st.columns(3)
                with col1:
                    render_kpi("Ventas Totales", f"${total_ventas/1_000_000:.1f} M", f"{delta_avance:+.1f}% vs Meta", "normal" if delta_avance >= 0 else "inverse")
                with col2:
                    render_kpi("Meta de Ventas", f"${total_meta/1_000_000:.1f} M")
                with col3:
                    card_color = "green" if avg_avance >= 100 else ("red" if avg_avance < 80 else "")
                    render_kpi("Avance General", f"{avg_avance:.1f}%", card_color=card_color)

                st.divider()
                
                # Gráficos
                graf_c1, graf_c2 = st.columns(2)
                with graf_c1:
                    fig_bar = px.bar(
                        df_filtrada, x='Vendedor', y=['Ventas ($)', 'Meta ($)'], barmode='group',
                        title='Rendimiento de Ventas vs. Meta (Filtrado)',
                        color_discrete_map={'Ventas ($)': COLOR_SECUNDARIO, 'Meta ($)': COLOR_ACENTO_ROJO}
                    )
                    st.plotly_chart(fig_bar, use_container_width=True)
                    
                    fig_donut = px.pie(
                        SAMPLE_DATA['ventas_categoria'], values='Ventas ($)', names='Categoria', 
                        title='Ventas por Categoría (Global)', hole=0.4,
                        color_discrete_sequence=px.colors.sequential.Blues_r
                    )
                    st.plotly_chart(fig_donut, use_container_width=True)

                with graf_c2:
                    fig_line = px.line(
                        SAMPLE_DATA['ventas_dia'], x='Fecha', y='Ventas ($)',
                        title='Tendencia de Ventas (Últimos 30 días)',
                        markers=True
                    )
                    fig_line.update_traces(line_color=COLOR_PRIMARIO)
                    st.plotly_chart(fig_line, use_container_width=True)

    # --- DEMO 2: ASISTENTE PROACTIVO (SIMULACIÓN IA) ---
    with tab2:
        st.subheader("Asistente Proactivo (Simulación de Análisis IA)")
        
        with st.container(border=True):
            vendedor_seleccionado = st.selectbox(
                "Simular plan de acción para:", 
                options=SAMPLE_DATA['ventas_vendedor']['Vendedor'].unique(),
                index=0
            )
            
            if st.button("Generar Plan de Acción (Simulación IA)", type="primary", use_container_width=True):
                with st.spinner(f"Analizando historial de {vendedor_seleccionado}... (Simulación)"):
                    time.sleep(1.5)
                
                if vendedor_seleccionado == 'DIEGO GARCIA':
                    st.info(
                        "**🎯 Oportunidad de Venta Cruzada (Demo):**\n"
                        "Tus clientes **'Ferretería El Tornillo'** y **'Depósito Central'** compran 'Discos de Corte' (A-101) pero "
                        "nunca han comprado 'Gafas de Seguridad' (D-404). ¡Ofréceles en su próximo pedido!"
                    )
                    st.warning(
                        "**🔥 Cliente en Riesgo (Análisis RFM - Demo):**\n"
                        "Tu cliente **'Construcciones Andinas'** no ha comprado en 62 días (Segmento: 'En Riesgo'). "
                        "Su frecuencia de compra habitual es de 30 días. **¡Acción: Llama hoy!**"
                    )
                elif vendedor_seleccionado == 'ANGELA CONTRERAS':
                    st.success(
                        "**💎 Cliente Campeón (Análisis RFM - Demo):**\n"
                        "Tu cliente **'Maestro SAS'** es un 'Campeón' (Compra reciente, frecuente y de alto valor). "
                        "**Acción: Agradécele** y ofrécele un producto nuevo de nicho."
                    )
                    st.info(
                        "**🎯 Oportunidad de Venta Cruzada (Demo):**\n"
                        "El cliente **'Decorar SAS'** ha comprado 'Guantes' (E-505) 3 veces. "
                        "El 80% de clientes que compran guantes, también compran 'Gafas de Seguridad' (D-404). **¡Acción: Ofrecer!**"
                    )
                else:
                    st.success(
                        "**🚀 Cliente Nuevo con Alto Potencial (Demo):**\n"
                        "El cliente **'Inversiones ABC'** (asignado a ti) acaba de hacer su primera compra por un valor alto. "
                        "**Acción: Llama para fidelizarlo** y asegurar una segunda compra."
                    )

    # --- DEMO 3: CATÁLOGO Y COTIZADOR (INTERACTIVO) ---
    with tab3:
        st.subheader("Catálogo Interactivo y Cotizador Profesional")
        st.markdown("Transformamos sus listas de precios estáticas en una herramienta de ventas. **Pruebe 'Añadir al Carrito'** y genere una cotización PDF con los productos que seleccionó.")
        
        with st.container(border=True):
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.subheader("Catálogo de Productos")
                catalogo_df = SAMPLE_DATA['catalogo_productos']
                
                num_cols = 3
                product_cols = st.columns(num_cols)
                
                for i, row in catalogo_df.iterrows():
                    col = product_cols[i % num_cols]
                    with col:
                        with st.container(border=True):
                            st.image(row['ImagenURL'], use_column_width=True)
                            st.markdown(f"**{row['Producto']}**")
                            st.markdown(f"<span style='color: {COLOR_PRIMARIO}; font-size: 1.1rem; font-weight: bold;'>${row['Vlr. Unitario']:,.0f}</span>", unsafe_allow_html=True)
                            st.markdown(f"<span style='color: {COLOR_TEXTO_SECUNDARIO}; font-size: 0.9rem;'>Stock: {row['Stock']}</span>", unsafe_allow_html=True)
                            
                            # --- CORRECCIÓN (BUG FIX 6.8) ---
                            # 'use_column_width' es incorrecto para st.button. Se cambia a 'use_container_width'.
                            if st.button(f"🛒 Añadir", key=f"add_cart_{row['Referencia']}", use_container_width=True):
                                if row['Referencia'] in st.session_state.cart['Referencia'].values:
                                    st.toast(f"'{row['Producto']}' ya está en el carrito.", icon="⚠️")
                                else:
                                    new_item = pd.DataFrame({
                                        'Referencia': [row['Referencia']],
                                        'Producto': [row['Producto']],
                                        'Cantidad': [1],
                                        'Vlr. Unitario': [row['Vlr. Unitario']],
                                        'Total': [row['Vlr. Unitario']]
                                    })
                                    st.session_state.cart = pd.concat([st.session_state.cart, new_item], ignore_index=True)
                                    st.toast(f"'{row['Producto']}' añadido al carrito!", icon="✅")
                                    # --- CORRECCIÓN (BUG FIX 6.6) ---
                                    # st.rerun() # Esta línea se elimina. Causa el error 'NotFoundError: removeChild'.
                                    # El 'rerun' implícito de la actualización de session_state es suficiente.

            with col2:
                st.subheader("Cotización en Proceso")
                
                if st.session_state.cart.empty:
                    st.info("Su carrito de cotización está vacío. Añada productos del catálogo.")
                else:
                    st.markdown("Puede editar la **Cantidad** directamente en la tabla:")
                    
                    edited_cart = st.data_editor(
                        st.session_state.cart,
                        column_config={
                            "Cantidad": st.column_config.NumberColumn(min_value=1, step=1),
                            "Total": st.column_config.NumberColumn(format="$ %d", disabled=True),
                            "Vlr. Unitario": st.column_config.NumberColumn(format="$ %d", disabled=True),
                            "Referencia": st.column_config.Column(disabled=True),
                            "Producto": st.column_config.Column(disabled=True),
                        },
                        use_container_width=True, 
                        hide_index=True,
                        key="cart_editor"
                    )
                    
                    # Recalcular totales si el carrito cambió
                    edited_cart['Total'] = edited_cart['Cantidad'] * edited_cart['Vlr. Unitario']
                    if not edited_cart.equals(st.session_state.cart):
                        st.session_state.cart = edited_cart
                        st.rerun()

                    subtotal = st.session_state.cart['Total'].sum()
                    st.metric("Subtotal Cotización", f"${subtotal:,.0f}")
                    
                    pdf_data = generar_demo_pdf(
                        st.session_state.cart,
                        "Cotización de Ejemplo",
                        "Este es un ejemplo de cotización profesional generada automáticamente por el sistema de GM-DATOVATE con los productos seleccionados."
                    )
                    st.download_button(
                        label="📄 Descargar PDF Profesional (Demo)",
                        data=pdf_data,
                        file_name="Demo_Cotizacion_GM-DATOVATE.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        type="primary"
                    )
                    if st.button("Vaciar Carrito", use_container_width=True):
                        st.session_state.cart = pd.DataFrame(columns=['Referencia', 'Producto', 'Cantidad', 'Vlr. Unitario', 'Total'])
                        st.rerun()


def render_pagina_operaciones():
    """Demo de la Suite de Operaciones y Logística."""
    
    st.markdown(f"<h2 style='color: {COLOR_PRIMARIO};'>🏭 Operaciones y Logística</h2>", unsafe_allow_html=True)
    st.markdown("Automatización de la cadena de suministro, desde el proveedor hasta la bodega, con inteligencia de datos.")
    st.divider()
    
    tab1, tab2, tab3 = st.tabs([
        "Demo 1: Abastecimiento Inteligente (Interactivo)",
        "Demo 2: Control de Inventario Móvil",
        "Demo 3: Sincronización (ETL Simulado)"
    ])

    # --- DEMO 1: ABASTECIMIENTO INTELIGENTE (INTERACTIVO) ---
    with tab1:
        st.subheader("Tablero de Abastecimiento Inteligente")
        st.markdown("Este módulo calcula la **Necesidad Real** y genera un plan de acción. **Pruebe editar las celdas 'Stock (Total)' o 'Necesidad Real'** y vea cómo se actualizan las sugerencias al instante.")
        
        with st.container(border=True):
            st.subheader("Sugerencias de Abastecimiento (Editable)")
            
            df_abastecimiento = SAMPLE_DATA['sugerencia_abastecimiento'].copy()
            
            edited_df = st.data_editor(
                df_abastecimiento,
                column_config={
                    "SKU": st.column_config.Column(disabled=True),
                    "Producto": st.column_config.Column(width="large", disabled=True),
                    "Stock (Total)": st.column_config.NumberColumn(min_value=0, step=10),
                    "Stock Tránsito": st.column_config.NumberColumn(min_value=0, step=10),
                    "Necesidad Real": st.column_config.NumberColumn(min_value=0, step=10),
                    "Sugerencia Traslado": st.column_config.NumberColumn(format="%d", disabled=True),
                    "Sugerencia Compra": st.column_config.NumberColumn(format="%d", disabled=True)
                },
                use_container_width=True, 
                hide_index=True,
                key="abastecimiento_editor"
            )

            # --- Lógica de recálculo sobre el DF editado ---
            edited_df['Sugerencia Traslado'] = 0 
            edited_df['Sugerencia Compra'] = (edited_df['Necesidad Real'] - (edited_df['Stock (Total)'] + edited_df['Stock Tránsito'])).clip(lower=0).astype(int)
            
            # --- CORRECCIÓN DE DEPRECACIÓN: .applymap -> .map ---
            st.dataframe(
                edited_df.style
                    .map(lambda x: f'background-color: {COLOR_ACENTO_ROJO}; color: white; font-weight: bold;' if x > 0 else '', subset=['Sugerencia Compra'])
                    .format({"Stock (Total)": "{:,.0f}", "Stock Tránsito": "{:,.0f}", "Necesidad Real": "{:,.0f}", "Sugerencia Traslado": "{:,.0f}", "Sugerencia Compra": "{:,.0f}"}),
                use_container_width=True, hide_index=True
            )
            
            # --- Botones de acción (Usa la función corregida) ---
            st.subheader("Generación de Órdenes (Demo)")
            df_orden = edited_df[edited_df['Sugerencia Compra'] > 0]
            df_traslado = edited_df[edited_df['Sugerencia Traslado'] > 0]
            
            pdf_data = generar_demo_pdf(
                df_orden[['SKU', 'Producto', 'Sugerencia Compra']],
                "Orden de Compra (Demo)",
                "Documento de ejemplo generado automáticamente para el proveedor, basado en las sugerencias del sistema."
            )
            
            excel_data = generar_demo_excel({
                "Orden_de_Compra": df_orden,
                "Detalle_Traslados": df_traslado,
                "Reporte_Completo": edited_df
            })

            c1, c2 = st.columns(2)
            c1.download_button(
                label="📄 Descargar Orden de Compra PDF (Demo)", data=pdf_data,
                file_name="Demo_Orden_de_Compra.pdf", mime="application/pdf", use_container_width=True, type="primary",
                disabled=df_orden.empty
            )
            c2.download_button(
                label="📊 Descargar Reporte Excel (Demo)", data=excel_data,
                file_name="Demo_Reporte_Abastecimiento.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True,
                disabled=edited_df.empty or excel_data is None # Deshabilitar si falla la generación
            )

    # --- DEMO 2: CONTROL DE INVENTARIO MÓVIL ---
    with tab2:
        st.subheader("Aplicación Móvil de Conteo Físico")
        st.markdown("Digitalizamos el conteo en bodega. El gerente asigna tareas y el operario las ejecuta en una app móvil con escáner y conteo parcial.")
        
        with st.container(border=True):
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("Vista del Operario (Móvil)")
                st.text_input("Buscar por Escáner o Referencia:", "770123456789", key="demo_scan")
                st.success("Producto Encontrado: 'Gafa de Seguridad'")
                render_kpi("Stock Teórico", 120, card_color="red")
                
                st.markdown("**Conteo Parcial (Calculadora):**")
                qty = st.number_input("Añadir cantidad:", value=0, step=1, key="demo_qty")
                if st.button("Registrar Cantidad", type="primary", use_container_width=True):
                    st.toast(f"Se registraron {qty} unidades. El total se actualizará.", icon="✅")

            with col2:
                st.subheader("Resumen de Conteo del Operario")
                st.dataframe(pd.DataFrame({
                    'Producto': ['Gafa de Seguridad', 'Disco de Corte'],
                    'Teórico': [120, 500],
                    'Contado': [118, 505],
                    'Historial Conteo': ["+50, +50, +10, +8", "+500, +5"],
                    'Diferencia': [-2, 5]
                }), use_container_width=True, hide_index=True)
                if st.button("Enviar Conteo Final para Revisión (Demo)", use_container_width=True):
                    st.success("¡Conteo enviado! El gerente será notificado.")
                    st.balloons()
            
    # --- DEMO 3: SINCRONIZACIÓN (ETL SIMULADO) ---
    with tab3:
        st.subheader("Sincronización Maestra de Inventario (ETL)")
        st.markdown("Desarrollamos un proceso que se conecta a su ERP (vía Dropbox, FTP, etc.), lee los archivos, los transforma y actualiza la base de datos central en la nube.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.graphviz_chart("""
                digraph "ETL Process" {
                    node [shape=box, style="filled,rounded", fontname="Arial", fontsize=12];
                    graph [bgcolor="transparent"];

                    erp [label="1. ERP Exporta CSV/XLSX\n(Ej: 'Rotacion.csv')", shape=cylinder, fillcolor="#E3F2FD"];
                    script [label="2. Script de Sincronización\n(Python + Pandas)", shape=component, fillcolor="#D1C4E9"];
                    nube [label="3. Base de Datos Maestra\n(SQL / Google Sheets)", shape=cylinder, fillcolor="#C8E6C9"];
                    apps [label="4. Todo el Ecosistema\n(Cotizador, BI, App Móvil)", shape=display, fillcolor="#FFF9C4"];

                    erp -> script [label=" Lee y Transforma"];
                    script -> nube [label=" Actualiza y Añade Nuevos"];
                    nube -> apps [label=" Alimenta Datos Vivos"];
                }
            """)

        with col2:
            st.subheader("Simulación de Carga ETL")
            st.markdown("Suba un archivo CSV o Excel (de ejemplo) para simular cómo el sistema lo procesa, identifica cambios y actualiza la base de datos.")
            
            uploaded_file = st.file_uploader("Subir archivo de inventario (CSV/Excel)", type=["csv", "xlsx"])
            
            if uploaded_file:
                try:
                    if uploaded_file.name.endswith('.csv'):
                        df_nuevo = pd.read_csv(uploaded_file)
                    else:
                        df_nuevo = pd.read_excel(uploaded_file)
                    
                    st.success("¡Archivo cargado y procesado con éxito!")
                    with st.spinner("Comparando con la base de datos maestra... (Simulación)"):
                        time.sleep(1)
                        st.subheader("Reporte de Sincronización (Demo)")
                        
                        st.info("Resumen de Cambios:\n* **2** productos actualizados (Stock y Precio).\n* **1** producto nuevo detectado ('F-606').")
                        st.markdown("**Productos Nuevos Detectados:**")
                        st.dataframe(pd.DataFrame({'SKU': ['F-606'], 'Producto': ['Lija de Agua'], 'Stock': [1000]}), use_container_width=True, hide_index=True)
                        st.markdown("**Muestra de Datos Cargados:**")
                        st.dataframe(df_nuevo.head(3), use_container_width=True)

                except Exception as e:
                    st.error(f"Error al procesar el archivo: {e}")


def render_pagina_finanzas():
    """Demo de la Suite Financiera."""
    
    st.markdown(f"<h2 style='color: {COLOR_PRIMARIO};'>🏦 Finanzas y Tesorería</h2>", unsafe_allow_html=True)
    st.markdown("Controle el flujo de caja, automatice la contabilidad y gestione el riesgo de cartera como nunca antes.")
    st.divider()
    
    tab1, tab2, tab3 = st.tabs([
        "Demo 1: Dashboard de Cartera (AR)",
        "Demo 2: Automatización Contable (Interactivo)",
        "Demo 3: Automatización de Riesgo"
    ])

    with tab1:
        st.subheader("Dashboard de Gestión de Cartera (AR)")
        
        with st.container(border=True):
            df_cartera = SAMPLE_DATA['cartera_antiguedad']
            total_cartera = df_cartera['Valor ($)'].sum()
            total_vencido = df_cartera[df_cartera['Rango'] != 'Al día']['Valor ($)'].sum()
            porc_vencido = (total_vencido / total_cartera * 100) if total_cartera > 0 else 0

            kpi1, kpi2, kpi3 = st.columns(3)
            with kpi1:
                render_kpi("Cartera Total", f"${total_cartera/1_000_000:.1f} M")
            with kpi2:
                render_kpi("Cartera Vencida", f"${total_vencido/1_000_000:.1f} M", card_color="red")
            with kpi3:
                render_kpi("Índice de Morosidad", f"{porc_vencido:.1f}%", card_color="red")

            fig_pie = px.pie(
                df_cartera, values='Valor ($)', names='Rango', title='Deuda por Antigüedad',
                hole=0.4, color='Rango', color_discrete_map=dict(zip(df_cartera['Rango'], df_cartera['Color']))
            )
            st.plotly_chart(fig_pie, use_container_width=True)
            
            st.divider()
            
            st.subheader("Gestión de Cliente Individual (Demo Interactiva)")
            df_detalle = SAMPLE_DATA['cartera_detalle']
            cliente_demo_nombre = st.selectbox("Seleccione un cliente para gestionar:", df_detalle['Cliente'].unique())
            
            if cliente_demo_nombre:
                cliente_demo_data = df_detalle[df_detalle['Cliente'] == cliente_demo_nombre].copy() # Usar copia para evitar SettingWithCopyWarning
                # Se asegura de que se selecciona la primera fila si hay múltiples facturas del mismo cliente para la info de contacto
                cliente_info = cliente_demo_data.iloc[0] 
                
                st.dataframe(cliente_demo_data[['Factura', 'Fecha Vencimiento', 'Días Vencido', 'Monto']].style.format({
                    "Días Vencido": "{:,.0f}",
                    "Monto": "${:,.0f}"
                }), use_container_width=True, hide_index=True)
                
                mensaje_wa = (
                    f"👋 ¡Hola {cliente_info['Cliente']}! Te saludamos desde GM-DATOVATE (Demo).\n\n"
                    f"Te recordamos tu saldo vencido. Tu factura *{cliente_info['Factura']}* por *${cliente_info['Monto']:,.0f}* "
                    f"presenta *{cliente_info['Días Vencido']} días* de vencimiento.\n\n"
                    f"Agradecemos tu pronta gestión."
                )
                url_wa = f"https://wa.me/{cliente_info['Teléfono']}?text={urllib.parse.quote(mensaje_wa)}"

                c1, c2, c3 = st.columns(3)
                
                # Esta es la línea que genera el error si falta la dependencia del sistema
                pdf_cartera = generar_demo_pdf_cartera(cliente_demo_data, cliente_info, fig_pie)
                
                c1.download_button(
                    label="📄 Descargar PDF (Demo)", data=pdf_cartera,
                    file_name=f"Cartera_{cliente_info['Cliente']}.pdf", mime="application/pdf",
                    use_container_width=True
                )
                
                if c2.button("✉️ Enviar Email (Demo)", use_container_width=True):
                    st.toast("Simulación: Email enviado a " + cliente_info['Email'], icon="✉️")

                c3.link_button("📲 Enviar WhatsApp (Demo)", url_wa, use_container_width=True)

    # --- DEMO 2: AUTOMATIZACIÓN CONTABLE (INTERACTIVO) ---
    with tab2:
        st.subheader("Automatización Contable (Cuadres y Recibos)")
        st.markdown("Eliminamos la digitación manual. Las tiendas llenan un formulario digital. **Pruebe editar los valores** y vea cómo se calcula la diferencia en tiempo real.")

        with st.container(border=True):
            st.subheader("Simulación de Cuadre de Caja Digital")
            c1, c2 = st.columns(2)
            c1.text_input("Tienda", "Armenia", disabled=True, key="demo_tienda_2")
            c2.date_input("Fecha", datetime.now().date(), disabled=True, key="demo_fecha_2")
            
            venta_total_sistema = 5_000_000
            render_kpi("Venta Total (Sistema)", f"${venta_total_sistema:,.0f}")
            
            st.markdown("##### Desglose de Pagos (Editable)")
            
            edited_cuadre = st.data_editor(
                st.session_state.cuadre_data,
                column_config={
                    "Tipo": st.column_config.Column(disabled=True),
                    "Valor": st.column_config.NumberColumn(format="$ %d", min_value=0, step=10000)
                },
                use_container_width=True,
                hide_index=True,
                key="cuadre_editor"
            )
            
            if not edited_cuadre.equals(st.session_state.cuadre_data):
                st.session_state.cuadre_data = edited_cuadre
                st.rerun()
            
            total_desglose = st.session_state.cuadre_data['Valor'].sum()
            diferencia = venta_total_sistema - total_desglose

            c_kpi1, c_kpi2 = st.columns(2)
            with c_kpi1:
                render_kpi("Total Desglose", f"${total_desglose:,.0f}")
            with c_kpi2:
                if diferencia == 0:
                    render_kpi("DIFERENCIA", f"${diferencia:,.0f}", delta="CUADRE PERFECTO", delta_color="off", card_color="green")
                else:
                    render_kpi("DIFERENCIA", f"${diferencia:,.0f}", delta="DESCUADRADO", delta_color="normal", card_color="red")
            
            demo_txt = "FECHA|CONSECUTIVO|CUENTA|...|DEBITO|CREDITO\n2025-11-05|1001|111005|...|2500000|0\n2025-11-05|1001|413501|...|0|2500000\n"
            st.download_button(
                label="💾 Descargar .TXT para ERP (Demo)", data=demo_txt,
                file_name="Demo_Contable_GM-DATOVATE.txt", mime="text/plain",
                use_container_width=True, type="primary", disabled=(diferencia != 0)
            )

    # --- DEMO 3: AUTOMATIZACIÓN DE RIESGO ---
    with tab3:
        st.subheader("Automatización de Riesgo (Integración Externa)")
        st.markdown("El sistema cruza automáticamente nuestra cartera con reportes de agencias externas (como Covinoc). Identifica discrepancias y genera los archivos de acción masiva.")
        
        with st.container(border=True):
            st.subheader("Resultados del Cruce Automático (Demo)")
            
            excel_demo_data = generar_demo_excel({
                "1_Facturas_a_Subir": SAMPLE_DATA['covinoc_subir'],
                "2_Facturas_a_Exonerar": SAMPLE_DATA['covinoc_exonerar']
            })
            
            st.warning("**Acción: Estas facturas están en nuestra cartera pero no en la agencia. Deben subirse.**")
            st.dataframe(SAMPLE_DATA['covinoc_subir'], use_container_width=True, hide_index=True)
            
            st.success("**Acción: Estas facturas ya fueron pagadas (no están en cartera) pero siguen activas en la agencia. Deben exonerarse.**")
            st.dataframe(SAMPLE_DATA['covinoc_exonerar'], use_container_width=True, hide_index=True)

            st.download_button(
                "📥 Descargar Reporte de Acciones (Excel Demo)", 
                excel_demo_data, 
                file_name="Demo_Reporte_Riesgo_GM-DATOVATE.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                use_container_width=True, type="primary"
            )

def render_pagina_integracion():
    """Demo de la Suite de Integración y Futuro (IA)."""
    
    st.markdown(f"<h2 style='color: {COLOR_PRIMARIO};'>🤖 Integración y Futuro (IA)</h2>", unsafe_allow_html=True)
    st.markdown("Conectamos todos los procesos, desde la vinculación de un cliente hasta el servicio post-venta con IA.")
    st.divider()

    tab1, tab2 = st.tabs([
        "Demo 1: Portal de Vinculación Digital (Interactivo)",
        "Demo 2: Agente IA (Chatbot Interactivo)"
    ])

    # --- DEMO 1: PORTAL DE VINCULACIÓN (INTERACTIVO) ---
    with tab1:
        st.subheader("Portal de Vinculación Digital de Clientes")
        st.markdown("Un portal público para que sus nuevos clientes se registren. El sistema captura datos, obtiene **firma digital** y valida identidad con **código OTP**. **Pruebe el flujo de OTP.**")
        
        with st.container(border=True):
            col1, col2 = st.columns(2)
            with col1:
                st.text_input("Razón Social*", "Mi Empresa S.A.S.", key="demo_rs")
                st.text_input("NIT*", "900.123.456-7", key="demo_nit")
                st.text_input("Email de Facturación*", "pagos@miempresa.com", key="demo_email_fact")
                email_otp = st.text_input("Email de Notificaciones (para OTP)*", "gerencia@miempresa.com", key="demo_email_otp")
            with col2:
                st.text_input("Representante Legal*", "Juan Pérez", key="demo_rl")
                st.text_input("C.C. del Representante*", "1.234.567.890", key="demo_cc")
                st.info("Por favor, firme en el recuadro:")
                
                # Canvas de firma (responsivo gracias al CSS inyectado)
                st_canvas(
                    stroke_width=3, stroke_color="#000000",
                    background_color="#FFFFFF", height=130, width=400,
                    key="canvas_demo"
                )
            
            st.divider()
            
            # --- Flujo de OTP ---
            st.subheader("Validación de Identidad (OTP)")
            otp_c1, otp_c2 = st.columns([1, 2])
            
            with otp_c1:
                if st.button("Enviar Código OTP (Simulación)", use_container_width=True, disabled=st.session_state.otp_sent):
                    with st.spinner(f"Enviando código a {email_otp}..."):
                        st.session_state.otp_code = str(np.random.randint(100000, 999999))
                        time.sleep(1)
                        st.session_state.otp_sent = True
                        st.success(f"¡Código enviado! (Simulación: es {st.session_state.otp_code})")
                        st.rerun()

            with otp_c2:
                otp_input = st.text_input("Código OTP enviado a su email", "******", max_chars=6, key="demo_otp", disabled=not st.session_state.otp_sent)
            
            otp_validado = (otp_input == st.session_state.otp_code) and st.session_state.otp_sent

            if st.session_state.otp_sent and not otp_validado and len(otp_input) == 6:
                st.warning("Código OTP incorrecto. Intente de nuevo.")

            if st.button("Finalizar Vinculación y Generar PDF (Demo)", use_container_width=True, type="primary", disabled=not otp_validado):
                st.success("¡Vinculación Simulada! En una implementación real, esto generaría un PDF legal y lo archivaría en la nube.")
                st.balloons()
                # Resetear OTP
                st.session_state.otp_sent = False
                st.session_state.otp_code = ""
                st.rerun() # Opcional: recargar para resetear el formulario

    # --- DEMO 2: AGENTE IA (CHATBOT INTERACTIVO) ---
    with tab2:
        st.subheader("El Agente IA (Chatbot de WhatsApp)")
        st.markdown("Un Chatbot con IA conectado a su ecosistema de datos. **Pruebe a preguntarle algo** como 'hola', 'cuál es mi deuda' o 'tienen stock de disco de corte'.")
        
        with st.container(border=True):
            st.subheader("Simulación de Chat (WhatsApp)")

            chat_container = st.container(height=400)

            for message in st.session_state.chat_messages:
                with chat_container.chat_message(message["role"]):
                    st.markdown(message["content"])
            
            def get_bot_response(user_message):
                time.sleep(1)
                low_msg = user_message.lower()
                
                if "deuda" in low_msg or "cartera" in low_msg or "factura" in low_msg:
                    return (
                        "¡Hola! Soy **DATO** 🕵️‍♂️. Consulté tu estado de cuenta (simulación) y veo lo siguiente:\n\n"
                        "1.  **Estado de Cartera:** Tienes una deuda vencida de **$1,200,000**.\n"
                        "2.  **Factura Vencida:** La factura FV-901 por $1,200,000 tiene 45 días de vencimiento.\n\n"
                        "¿Te gustaría que te envíe el estado de cuenta a tu correo?"
                    )
                elif "stock" in low_msg or "inventario" in low_msg or "disco" in low_msg or "tornillo" in low_msg:
                    return (
                        "¡Claro! Consulté nuestro inventario en tiempo real (simulación):\n\n"
                        "* **'Disco Corte 4-1/2\"' (Ref: A-101):**\n"
                        "    * Bodega CEDI: **500** unidades\n"
                        "* **'Tornillo Drywall 6x1' (Ref: B-202):**\n"
                        "    * Bodega CEDI: **15,000** unidades\n"
                        "    * En Tránsito: **5,000** unidades\n\n"
                        "¿Deseas que un vendedor te contacte para hacer un pedido?"
                    )
                elif "hola" in low_msg or "buenos dias" in low_msg:
                    return (
                        "¡Hola! Soy **DATO**, tu asistente de IA en **GM-DATOVATE** (en modo demo). \n\n"
                        "Estoy conectado a los datos de Ventas, Cartera e Inventario. \n\n"
                        "Puedes preguntarme cosas como:\n"
                        "* '¿Cuál es mi deuda?'\n"
                        "* '¿Tienen stock de tornillos?'"
                    )
                else:
                    return (
                        "Soy una demo de IA. No puedo procesar esa solicitud específica, "
                        "pero estoy diseñado para responder a consultas sobre **deuda**, **stock** y **precios**. "
                        "¡Intenta con una de esas!"
                    )

            if prompt := st.chat_input("Escribe tu consulta aquí..."):
                st.session_state.chat_messages.append({"role": "user", "content": prompt})
                with chat_container.chat_message("user"):
                    st.markdown(prompt)
                
                with chat_container.chat_message("assistant"):
                    with st.spinner("DATO está pensando... 🤖"):
                        response = get_bot_response(prompt)
                        st.markdown(response)
                
                st.session_state.chat_messages.append({"role": "assistant", "content": response})


def render_pagina_contacto():
    """Renderiza la página final de Contacto / CTA."""
    
    st.markdown('<div id="contacto"></div>', unsafe_allow_html=True)
    
    with st.container():
        st.markdown('<div class="section-container contact-form-container">', unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown(f"""
            <h2 style="border: none; color: {COLOR_PRIMARIO};">Hablemos de su Transformación Digital 🚀</h2>
            <p style="font-size: 1.1rem; color: {COLOR_TEXTO_SECUNDARIO};">
                Lo que ha visto en este portafolio no es una teoría; es una 
                <strong>demostración de sistemas reales, funcionales e interactivos</strong>
                que ya están generando un valor incalculable.
                <br><br>
                ¿Está listo para dejar de "administrar" su negocio y empezar a "dirigirlo"? 
                Complete el formulario y uno de nuestros arquitectos de soluciones 
                se pondrá en contacto con usted.
            </p>
            """, unsafe_allow_html=True)
        
        with col2:
            with st.form(key="contact_form"):
                st.subheader("Iniciemos la Conversación")
                c1, c2 = st.columns(2)
                nombre = c1.text_input("Su Nombre*")
                empresa = c2.text_input("Su Empresa*")
                email = c1.text_input("Su Correo Electrónico*")
                telefono = c2.text_input("Su Teléfono/WhatsApp")
                
                desafio = st.text_area(
                    "¿Cuál es su mayor desafío operativo o de datos en este momento?*",
                    placeholder="Ej: 'Mi inventario nunca cuadra', 'Mi equipo de ventas es muy lento para cotizar', 'No tengo idea de mi cartera vencida en tiempo real'..."
                )
                
                submit = st.form_submit_button("Solicitar Consulta Estratégica", use_container_width=True)
                
                if submit:
                    if not all([nombre, empresa, email, desafio]):
                        st.warning("Por favor, complete todos los campos marcados con *.")
                    else:
                        st.success(f"¡Gracias, {nombre}! He recibido su solicitud. Nos pondremos en contacto con usted en {email} muy pronto.")
                        st.balloons()
                        
        st.markdown('</div>', unsafe_allow_html=True)

# ======================================================================================
# --- NAVEGACIÓN PRINCIPAL (Flujo de Página Única) ---
# ======================================================================================

# 1. Renderiza la página de "Inicio" (Hero + Resumen + Equipo)
render_pagina_inicio()

# 2. Título para la sección de demos
st.markdown(f"<h2 style='text-align: center; border: none; margin-top: 4rem; margin-bottom: 0rem; padding-bottom: 0; color: {COLOR_PRIMARIO};'>Explore las Demos Interactivas</h2>", unsafe_allow_html=True)

# 3. Pestañas (Tabs) para las demos interactivas
tab_com, tab_ops, tab_fin, tab_int = st.tabs([
    "🧠 Inteligencia Comercial",
    "🏭 Operaciones y Logística",
    "🏦 Finanzas y Tesorería",
    "🤖 Integración y Futuro (IA)"
])

with tab_com:
    render_pagina_comercial()

with tab_ops:
    render_pagina_operaciones()

with tab_fin:
    render_pagina_finanzas()

with tab_int:
    render_pagina_integracion()

# 4. Renderiza el formulario de contacto al final de la página.
st.divider()
render_pagina_contacto()

# --- Footer Personalizado ---
st.markdown(f"""
<div style="background-color: {COLOR_PRIMARIO}; color: {COLOR_FONDO_SECUNDARIO}; padding: 2.5rem; text-align: center; margin-top: 3rem;">
    <p style="color: white; margin: 0; font-size: 1.1rem;">
        <strong>GM-DATOVATE</strong>
    </p>
    <p style="color: {COLOR_FONDO_SECUNDARIO}; margin: 0.5rem 0; font-size: 0.9rem;">
        Transformando Datos en Decisiones Estratégicas.
        <br>
        © {datetime.now().year} Todos los derechos reservados.
    </p>
</div>
""", unsafe_allow_html=True)
